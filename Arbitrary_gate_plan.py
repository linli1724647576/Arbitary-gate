import openpyxl


def extraction_digital(string):
    if string == '1米6':
        return 160
    d = "".join(list(filter(str.isdigit, string)))
    if d:
        return int(d[:3])
    else:
        return 0


all_data = []
excel = openpyxl.load_workbook(f'./3.xlsx')
sheet = excel.worksheets[0]
for row in sheet.iter_rows():
    if row[0].value is None:
        break
    all_data.append(row)
all_data = all_data[1:]

all_person = list()

for data in all_data:
    one_person = dict()
    one_person['id'] = data[0].value
    one_person['sex'] = data[7].value
    one_person['phone_number'] = data[9].value
    one_person['wx'] = data[10].value
    one_person['student_id'] = data[11].value
    one_person['college'] = data[12].value
    one_person['Professional_grade'] = data[13].value
    one_person['personal'] = data[14].value
    one_person['adjust'] = data[16].value
    one_person['personal_information'] = []
    one_person['match_tendency'] = []
    # 11、个人年龄
    # 12、身高
    # 13、所在校区
    # 14、你的专业是社科类还是理工类
    # 15、用一个词形容你的性格特征
    # 16、你的星座
    if isinstance(data[17].value, int):
        one_person['personal_information'].append(data[17].value)
    else:
        one_person['personal_information'].append(extraction_digital(data[17].value))
    if isinstance(data[18].value, int):
        one_person['personal_information'].append(data[18].value)
    else:
        one_person['personal_information'].append(extraction_digital(data[18].value))
    one_person['personal_information'].append(data[19].value)
    one_person['personal_information'].append(data[20].value)
    one_person['personal_information'].append(data[21].value)
    one_person['personal_information'].append(data[22].value)

    age = [28, 0]
    for i in range(3):
        if data[23 + i].value:
            if age[0] > 19 + (i * 3):
                age[0] = 19 + (i * 3)
            if age[1] < 21 + (i * 3):
                age[1] = 21 + (i * 3)
    if data[26].value:
        age[1] = 100
    one_person['match_tendency'].append(age)

    height = [181, 0]
    for i in range(6):
        if data[27 + i].value:
            if height[0] > 151 + (i * 5):
                height[0] = 151 + (i * 5)
            if height[1] < 155 + (i * 5):
                height[1] = 155 + (i * 5)
    if data[33].value:
        height[1] = 300
    if data[34].value:
        height = [0, 300]
    if height[0] == 151:
        height[0] = 150
    one_person['match_tendency'].append(height)

    if data[35].value == 3:
        one_person['match_tendency'].append([1, 2])
    else:
        one_person['match_tendency'].append([data[35].value])
    if data[36].value == 3:
        one_person['match_tendency'].append([1, 2])
    else:
        one_person['match_tendency'].append([data[36].value])

    if data[37].value == 9:
        one_person['match_tendency'].append(list(range(1, 9)))
    else:
        one_person['match_tendency'].append([data[37].value])

    if data[50].value:
        one_person['match_tendency'].append(list(range(1, 13)))
    else:
        constellation = []
        for i in range(1, 13):
            if data[37 + i].value:
                constellation.append(i)
        one_person['match_tendency'].append(constellation)

    pre_weight = [{1: 35, 2: 25, 3: 18, 4: 12, 5: 7, 6: 3},
                  {1: 35, 2: 25, 3: 18, 4: 12, 5: 7, 6: 3},
                  {1: 37, 2: 27, 3: 18, 4: 12, 5: 3, 6: 3},
                  {1: 40, 2: 30, 3: 19, 4: 3, 5: 3, 6: 3},
                  {1: 50, 2: 38, 3: 3, 4: 3, 5: 3, 6: 3},
                  {1: 65, 2: 7, 3: 7, 4: 7, 5: 7, 6: 7},
                  ]

    # 23、(身高)
    # 23、(校区)
    # 23、(年龄)
    # 23、(专业)
    # 23、(性格)
    # 23、(星座)

    # 11、个人年龄
    # 12、身高
    # 13、所在校区
    # 14、你的专业是社科类还是理工类
    # 15、用一个词形容你的性格特征
    # 16、你的星座
    select = []
    select_num = 0
    for i in range(1, 7):
        if data[50 + i].value != -2:
            select.append(data[50 + i].value)
            select_num += 1
        else:
            select.append(6)
    weight = []
    for x in select:
        weight.append(pre_weight[6 - select_num].get(x))
    one_person['weight'] = weight
    all_person.append(one_person)

education_background = ['硕士', '博士']
education_background_id = 2

personal = []
personal_id = []
for one_person in all_person:
    if one_person['personal'] and one_person['Professional_grade'] == education_background_id:
        personal.append(one_person)
        personal_id.append(one_person['id'])
print('二人组的有： ' + str(len(personal)) + '人')
boy = []
girl = []
for p in personal:
    if p['sex'] == 1:
        boy.append(p)
    else:
        girl.append(p)
print(f'二人组中男神 {len(boy)} 人 女神 {len(girl)}人')

map = [2, 0, 1, 3, 4, 5]

matrix_girl = dict()
for one_girl in girl:
    res = dict()
    match_tendency = one_girl['match_tendency']
    weight = one_girl['weight']
    for one_boy in boy:
        score = 0
        information = one_boy['personal_information']
        if match_tendency[0][0] <= information[0] <= match_tendency[0][1]:
            score += weight[2]
        if match_tendency[1][0] <= information[1] <= match_tendency[1][1]:
            score += weight[0]
        if information[2] in match_tendency[2]:
            score += weight[1]
        for i in range(3, 6):
            if information[i] in match_tendency[i]:
                score += weight[i]
        res[one_boy['id']] = score
    matrix_girl[one_girl['id']] = res

    print(one_girl['id'], res)

matrix_boy = dict()
for one_boy in boy:
    res = dict()
    match_tendency = one_boy['match_tendency']
    weight = one_boy['weight']
    for one_girl in girl:
        score = 0
        information = one_girl['personal_information']
        if match_tendency[0][0] <= information[0] <= match_tendency[0][1]:
            score += weight[2]
        if match_tendency[1][0] <= information[1] <= match_tendency[1][1]:
            score += weight[0]
        if information[2] in match_tendency[2]:
            score += weight[1]
        for i in range(3, 6):
            if information[i] in match_tendency[i]:
                score += weight[i]
        res[one_girl['id']] = score
    matrix_boy[one_boy['id']] = res
    print(one_boy['id'], res)

# score_integration
excel = openpyxl.Workbook()
sheet = excel.active

keys_girl = list(matrix_girl.keys())
keys_boy = list(matrix_boy.keys())

sheet.cell(1, 1).value = '备注：行代表女生，列代表男神，每个单元格有两个数，第一个数代表这个男生对于女生的分数，第二个数代表女神对于男生的分数。'

for i in range(len(keys_boy)):
    sheet.cell(1, i + 2).value = keys_boy[i]

for i in range(len(keys_girl)):
    sheet.cell(i + 2, 1).value = keys_girl[i]

for i in range(len(keys_girl)):
    for j in range(len(keys_boy)):
        sheet.cell(i + 2, j + 2).value = str(matrix_girl.get(keys_girl[i]).get(keys_boy[j])) + ',' + \
                                         str(matrix_boy.get(keys_boy[j]).get(keys_girl[i]))

excel.save(f'./Arbitrary_gate_plan/{education_background[education_background_id-1]}_score_integration.xlsx')

matched = []
lowest = 30
for i in range(5):
    for p in personal:
        if p['id'] in personal_id:
            if p['sex'] == 1:
                max_score = -1
                max_score_id = [-1, -1,]
                for k, v in matrix_boy.get(p['id']).items():
                    v1 = matrix_girl.get(k).get(p['id'])
                    if k in personal_id and v >= lowest and v1 >= lowest and v1+v > max_score:
                        max_score = v+v1
                        max_score_id = [k, p['id'], v1, v]

                if max_score > -1:
                    matched.append(max_score_id)
                    personal_id.remove(max_score_id[0])
                    personal_id.remove(max_score_id[1])
            else:
                max_score = -1
                max_score_id = [-1, -1]
                for k, v in matrix_girl.get(p['id']).items():
                    v1 = matrix_boy.get(k).get(p['id'])
                    if k in personal_id and v >= lowest and v1 >= lowest and v1 + v > max_score:
                        max_score = v + v1
                        max_score_id = [p['id'], k, v, v1]

                if max_score > -1:
                    matched.append(max_score_id)
                    personal_id.remove(max_score_id[0])
                    personal_id.remove(max_score_id[1])
    # lowest -= (i+1)*10

print(personal_id)
print(matched)

# score_integration
excel = openpyxl.Workbook()
sheet = excel.active

keys_girl = list(matrix_girl.keys())
keys_boy = list(matrix_boy.keys())

sheet.cell(1, 1).value = f'备注：二人组中男神 {len(boy)} 人，女神 {len(girl)}人，' \
                         f'匹配成功有{len(matched)}对'

match_fail = ",".join([str(x) for x in personal_id])
sheet.cell(2, 1).value = f'未匹配成功的有：{match_fail}，共{len(personal_id)}人'
sheet.cell(3, 1).value = 'girl'
sheet.cell(3, 2).value = 'boy'
sheet.cell(3, 3).value = '男生对于女生的得分'
sheet.cell(3, 4).value = '女生对于男生的得分'
for i in range(len(matched)):
    for j in range(4):
        sheet.cell(4 + i, 1 + j).value = matched[i][j]
excel.save(f'./Arbitrary_gate_plan/{education_background[education_background_id-1]}_match_{lowest}.xlsx')

