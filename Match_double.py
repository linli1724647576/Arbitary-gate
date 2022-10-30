from Person import Person
import openpyxl

class graphMatrix:
    def __init__(self):
        self.dict={}  # 结点对应矩阵下标的字典
        self.length=0  # 矩阵的长度
        self.matrix=[]

    # 添加节点
    def addVertex(self,key):
        self.dict[key]=self.length  # 键(顶点名称)->值(在矩阵中的下标)
        self.dict[self.length]=key  # 键(在矩阵中的下标)->值(顶点名称)

        lst=[]
        for i in range(self.length):  # 创建一行长度为self.length,全是0的列表
            lst.append(0)
        self.matrix.append(lst)

        for row in self.matrix:  # 为矩阵中每行添加新的一个单元，单元中值为0
            row.append(0)
        self.length += 1

    # 添加有向边
    def addDirectLine(self,start,ends,wt):
        startIndex=self.dict[start] # 起点在二维矩阵中的下标
        endIndex=self.dict[ends]  # 终点在二维矩阵中的下标
        self.matrix[startIndex][endIndex]=wt # 权值赋值


if __name__ == "__main__":
    all_data = []
    excel = openpyxl.load_workbook(f'./二次匹配人员-2人组（去掉校区要求）v2.xlsx')
    sheet = excel.worksheets[0]
    for row in sheet.iter_rows():
        if row[0].value is None:
            break
        all_data.append(row)
    all_data = all_data[1:]
    person_list = []
    person_dict = dict()  # Map 用于快速查找
    for data in all_data:
        p = Person(data)
        p.print_person()
        if p.to_or_eight=="个人":
            person_list.append(p)
            person_dict[p.id]=p


    boys = []
    girls = []
    for i in person_list:
        if i.sex == "男":
            boys.append(i)
        elif i.sex == "女":
            girls.append(i)

    matrix_girl = dict()
    for one_girl in girls:
        res = dict()
        for one_boy in boys:
            score = one_girl.get_score(one_boy)
            res[one_boy.id] = score
        matrix_girl[one_girl.id] = res
        # print(one_girl.id, res)

    matrix_boy = dict()
    for one_boy in boys:
        res = dict()
        for one_girl in girls:
            score = one_boy.get_score(one_girl)
            res[one_girl.id] = score
        matrix_boy[one_boy.id] = res
        # print(one_boy.id, res)

    # score_integration
    excel = openpyxl.Workbook()
    sheet = excel.active

    keys_girl = list(matrix_girl.keys())
    keys_boy = list(matrix_boy.keys())

    sheet.cell(1, 1).value = '备注：行代表女生，列代表男神，每个单元格有两个数，第一个数代表这个男生对于女生的分数，第二个数代表女神对于男生的分数。'

    for i in range(len(keys_boy)):
        sheet.cell(1, i + 2).value = keys_boy[i]

    for i in range(len(keys_girl)):
        sheet.cell(i + 2, 1).value = keys_girl[i]

    for i in range(len(keys_girl)):
        for j in range(len(keys_boy)):
            sheet.cell(i + 2, j + 2).value = str(matrix_girl.get(keys_girl[i]).get(keys_boy[j])) + ',' + \
                                             str(matrix_boy.get(keys_boy[j]).get(keys_girl[i]))

    excel.save(f'./Arbitrary_gate_plan/score_integration.xlsx')

    matched = []
    lowest = 200
    for i in range(5):
        for p in person_list:
            if p.sex== '男':
                max_score = -1
                max_score_id = [-1, -1, ]
                for k, v in matrix_boy.get(p.id).items():
                    # print("k=",k)  # 女生的序号
                    # print("v=",v)  # 男生对女生的分数
                    v1 = matrix_girl.get(k).get(p.id)   #女生对男生的分数
                    if person_dict[k] in person_list and v >= lowest and v1 >= lowest and v1 + v > max_score:
                        max_score = v + v1
                        max_score_id = [k, p.id, v1, v]   #[女生的序号，男生的序号，女生对男生的分数，男生对女生的分数]

                if max_score > -1:
                    matched.append(max_score_id)
                    male = person_dict[max_score_id[1]]
                    female = person_dict[max_score_id[0]]
                    if male in person_list:
                        person_list.remove(male)
                    if female in person_list:
                        person_list.remove(female)
            else:
                max_score = -1
                max_score_id = [-1, -1]
                for k, v in matrix_girl.get(p.id).items():
                    v1 = matrix_boy.get(k).get(p.id)
                    if person_dict[k] in person_list and v >= lowest and v1 >= lowest and v1 + v > max_score:
                        max_score = v + v1
                        max_score_id = [p.id, k, v, v1]   #[女生的序号，男生的序号，女生对男生的分数，男生对女生的分数]

                if max_score > -1:
                    matched.append(max_score_id)
                    male = person_dict[max_score_id[1]]
                    female = person_dict[max_score_id[0]]
                    if male in person_list:
                        person_list.remove(male)
                    if female in person_list:
                        person_list.remove(female)
        lowest -= (i+1)*10

    print(matched)
    print(len(matched))

    # score_integration
    excel = openpyxl.Workbook()
    sheet = excel.active

    keys_girl = list(matrix_girl.keys())
    keys_boy = list(matrix_boy.keys())

    sheet.cell(1, 1).value = f'备注：二人组中男神 {len(boys)} 人，女神 {len(girls)}人，' \
                             f'匹配成功有{len(matched)}对'

    match_fail = ",".join([str(x.id) for x in person_list])
    sheet.cell(2, 1).value = f'未匹配成功的有：{match_fail}，共{len(person_list)}人'
    sheet.cell(3, 1).value = 'girl'
    sheet.cell(3, 2).value = "姓名"
    sheet.cell(3, 3).value = "微信"
    sheet.cell(3, 4).value = 'boy'
    sheet.cell(3, 5).value = "姓名"
    sheet.cell(3, 6).value = "微信"
    sheet.cell(3, 7).value = '男生对于女生的得分'
    sheet.cell(3, 8).value = '女生对于男生的得分'
    for i in range(len(matched)):
        girl = person_dict[matched[i][0]]
        boy = person_dict[matched[i][1]]
        sheet.cell(4 + i, 1).value = girl.id
        sheet.cell(4 + i, 2).value = girl.name
        sheet.cell(4 + i, 3).value = girl.wx
        sheet.cell(4 + i, 4).value = boy.id
        sheet.cell(4 + i, 5).value = boy.name
        sheet.cell(4 + i, 6).value = boy.wx
        sheet.cell(4 + i, 7).value = matched[i][2]
        sheet.cell(4 + i, 8).value = matched[i][3]

        # for j in range(4):
        #     sheet.cell(4 + i, 1 + j).value = matched[i][j]
    excel.save(f'./Arbitrary_gate_plan/matched_result_8_again_11111111.xlsx')


    # gm = graphMatrix()  # 创建有向图
    # # 添加结点
    # for k in person_list:
    #     gm.addVertex(k)
    #
    # # 添加有向边
    # for i in range(len(person_list)):
    #     for j in range(i+1,len(person_list)):
    #         s = person_list[i].get_score(person_list[j])
    #         gm.addDirectLine(person_list[i],person_list[j],s)
    # print(gm.matrix)
